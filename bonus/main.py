import openpyxl
from statistics import mean
import matplotlib.pyplot as plt
from collections import  Counter
# from wordcloud import WordCloud

def average_salary(staff):
    all_salary = []
    for i in staff:
        all_salary.append(i[4])
    print('средняя зарплата %6.1f' % mean(all_salary))
    return mean(all_salary)


def salary_by_gender(staff):
    men_salary = []
    women_salary = []
    for i in staff:
        if i[1] == 'М':
            men_salary.append(i[4])
        else:
            women_salary.append(i[4])
    print('средняя зарплата среди мужчин %6.1f' % mean(men_salary))
    print('средняя зарплата среди женщин %6.1f' % mean(women_salary))
    return mean(men_salary), mean(women_salary)


def print_salary_by_position(staff):
    salary_by_position = {}
    for i in staff:
        tmp_position = i[2]
        if tmp_position in salary_by_position:
            salary_by_position[tmp_position].append(i[4])
        else:
            salary_by_position[tmp_position] = []
            salary_by_position[tmp_position].append(i[4])

    # Вывод всех значений средних зарплат
    # print(salary_by_position)

    # Сортировка по среднему значению зарплаты -> список
    salary_by_position = sorted(salary_by_position.items(), key=lambda kv: mean(kv[1]))

    for i in salary_by_position:
        print('средняя зарплата: %25s %6.1f ' % (i[0], mean(i[1])))


def show_statisctics(average_salary, ages, average_salary_by_age, average_men_salary, average_women_salary):
    # labels - подписи к данным, sizes - данные
    labels = 'Муж', 'Жен'
    sizes = [average_men_salary, average_women_salary]
    x = ages
    y1 = average_salary_by_age
    y2 = [average_salary for i in range(0, len(x))]
    # explode - это отступ (разрыв) от основных частей пирога.
    explode = (0.05, 0.05)

    fig = plt.figure()

    # 1 - show trend. 2 - show male/female division
    case = 1
    if case == 1:
        ax1 = fig.add_subplot(111)
        ax1.plot(x, y1, 'o--', color="blue", label="З\п")
        ax1.plot(x, y2, color="red", label="Средняя")
        ax1.legend(loc=4, bbox_to_anchor=(0.6, 0.0))
        ax1.set_xlabel('возраст', fontsize=10)
        ax1.grid()
        ax1.set_title('З\п по возрастам')
    elif case == 2:
        ax2 = fig.add_subplot(111)
        ax2.pie(sizes, explode=explode, labels=labels, autopct='%1.1f%%',
                shadow=True, startangle=90)
        ax2.axis('equal')
        ax2.set_title('З\п')
    # показать график
    plt.show()


def show_wordcloud(counter_obj):
    total_count = len(counter_obj)
    common = counter_obj.most_common(total_count)
    # wc = WordCloud(width=2600, height=2200, background_color="white", relative_scaling=1.0,
    #               collocations=False, min_font_size=10).generate_from_frequencies(dict(common))
    plt.axis("off")
    plt.figure(figsize=(9, 6))
    # plt.imshow(wc, interpolation="bilinear")
    plt.title("Должности")
    plt.xticks([])
    plt.yticks([])
    plt.tight_layout()
    file_name = 'tag_cloud.png'
    plt.savefig(file_name)
    plt.show()

def show_bar(counter_obj):
    heights = counter_obj.values()
    bars = counter_obj.keys()
    y_pos = range(len(bars))
    # Rotation of the bars names
    plt.bar(y_pos, heights, align='center', alpha=0.5)
    plt.xticks(y_pos, bars, rotation=90)
    plt.ylabel('Число вхождений')
    plt.title('Встречаемость должностей')
    file_name = 'tag_cloud.png'
    plt.savefig(file_name)
    plt.show()

def main_tags():
    input_file = './lab_04/Data.xlsx'
    staff = Counter()
    source_book = openpyxl.load_workbook(input_file)
    shield_names = source_book.sheetnames
    for shield_name in shield_names:
        data_shield = source_book[shield_name]
        rows = list(data_shield.rows)
        for row in rows:
            # row[0] - имя
            # row[1] - пол
            # row[2] - должность
            # row[3] - возраст
            # row[4] - зарплата
            if row[2].value is not None:
                staff[row[2].value] += 1
    # case = 1 - wordcloud lib, case = 2 - primary plot bar
    case = 2
    if case == 1:
        show_wordcloud(staff)
    elif case == 2:
        show_bar(staff)

def main_plot():
    input_file = './lab_04/Data.xlsx'
    staff = []

    average_salary_by_age = []
    source_book = openpyxl.load_workbook(input_file)
    shield_names = source_book.sheetnames
    for shield_name in shield_names:
        data_shield = source_book[shield_name]
        rows = list(data_shield.rows)
        rows.pop(0)
        member_salaries_by_age = []
        for row in rows:
            # row[0] - имя
            # row[1] - пол
            # row[2] - должность
            # row[3] - возраст
            # row[4] - зарплата
            tmp_member = (row[0].value, row[1].value, row[2].value, row[3].value, row[4].value)
            staff.append(tmp_member)
            member_salaries_by_age.append(tmp_member[4])
        print('возраст %8s лет -> средняя зарплата: %6.1f ' % (shield_name, mean(member_salaries_by_age)))
        average_salary_by_age.append(mean(member_salaries_by_age))

    total_average_salary = average_salary(staff)
    mean_men_salary, mean_women_salary = salary_by_gender(staff)
    print_salary_by_position(staff)
    show_statisctics(
        average_salary=total_average_salary,
        average_salary_by_age=average_salary_by_age,
        ages=shield_names,
        average_men_salary=mean_men_salary,
        average_women_salary=mean_women_salary
    )


if __name__ == '__main__':
    #main_plot()
    main_tags()