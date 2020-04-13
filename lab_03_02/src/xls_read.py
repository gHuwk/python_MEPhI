import openpyxl

book = openpyxl.load_workbook('sample.xlsx')

# show sheets names
print(book.sheetnames)

# choice of sheet
sheet = book['Data']

# get value
title_1 = sheet['A1'].value #access by  key of a worksheet
title_2 = sheet.cell(row=1, column=2) # access by row and column
pairs = sheet['B2':'C6']


# print values
print(title_1)
print(title_2.value)

for c1, c2 in pairs:
    print(c1.value, c2.value)

for row in sheet.iter_rows(min_row=2, max_col=3,max_row=6):
    for cell in row:
        size = ' '*(20-len(str(cell.value)))
        print(cell.value, end=size)
    print()

a = input()
