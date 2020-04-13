import openpyxl as XSL

book = XSL.Workbook()

# grab the active worksheet
sheet = book.active
ws = book.active

# You can create new worksheets using Workbook.create_sheet()
ws1 = book.create_sheet('Sheet 1') # insert at the end (default)
ws2 = book.create_sheet('Sheet 2', 2) # insert at first position

# you can change Title. color and so on
ws2.title = 'I fixed the title'

#insert data
ws1['A1'] = 1000 #access by  key of a worksheet
ws1['A2'] = 100
ws.cell(row=2, column=2).value = 100 # access by row and column

d = ws.cell(row = 3, column = 122, value = 100)
d.value = 322

ws.append((1,2)) # insert at the end of the workshhet
ws2.append([1,3,4])
# value should be a list, tuple, range of generator or dict

# you can create copy
target = book.copy_worksheet(ws)
target.title = ws.title + ' My Copy'

#save xlsx file
book.save('sample.xlsx')


# insert data
ws3 = book.create_sheet('Data')
ws3.append(['Town', 'distance, km', 'cost, RUB'])
data = (
    ('Anapa', 1000, 5000),
    ('Adler', 1500, 7000),
    ('Moscow', 400, 2000),
    ('Saint-Petersburg', 1000, 6000),
    ('Tula', 350, 400)
    )
for pair in data:
    ws3.append(pair)
book.save('sample.xlsx')



